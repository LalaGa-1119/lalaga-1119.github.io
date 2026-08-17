# -*- coding: utf-8 -*-
"""生成个人主页信息的 Excel 表格（产品经理版）"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ===== 样式 =====
title_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="38BDF8")
head_font = Font(name="微软雅黑", size=11, bold=True, color="0F172A")
head_fill = PatternFill("solid", fgColor="E0F2FE")
cell_font = Font(name="微软雅黑", size=11)
link_font = Font(name="微软雅黑", size=11, color="0563C1", underline="single")
thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_title(ws, ncols, title):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, title)
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 28

def write_cell(ws, r, c, value, alignment=left, font=cell_font):
    cell = ws.cell(r, c, value)
    cell.border = border
    cell.alignment = alignment
    cell.font = font
    if isinstance(value, str) and value.startswith("http"):
        cell.hyperlink = value
        cell.font = link_font
    return cell

# ===== Sheet 1: 基本信息 =====
ws = wb.active
ws.title = "基本信息"
data = [
    ["姓名", "高丽丽"],
    ["英文名", "Gao Lili"],
    ["性别", "女"],
    ["出生年月", "2000-06"],
    ["籍贯", "北京"],
    ["现居城市", "北京"],
    ["学历", "本科"],
    ["毕业院校", "某某大学"],
    ["专业", "计算机科学与技术"],
    ["手机", "138-0000-0000"],
    ["邮箱", "gaolili@example.com"],
    ["微信", "gaolili_wechat"],
    ["GitHub", "github.com/gaolili"],
    ["个人网站", "gaolili.me"],
    ["求职意向", "产品经理"],
    ["个人简介", "产品经理，现居北京，2 年产品经验，先后参与 6 个从 0 到 1 的项目，"
     "覆盖 B 端 SaaS 与移动端社区方向。擅长需求分析、原型设计与跨团队协作，"
     "坚持在「人人都是产品经理」等平台分享产品思考。"],
    ["代表文章", "《AI越主动体验就越好吗？从一篇成稿看排版工具的介入边界》"],
    ["文章链接", "https://www.woshipm.com/evaluating/6443589.html"],
]
style_title(ws, 2, "基本信息")
for r, (k, v) in enumerate(data, start=2):
    write_cell(ws, r, 1, k, alignment=center, font=head_font)
    write_cell(ws, r, 2, v)
for row in ws.iter_rows(min_row=2, max_row=len(data) + 1, max_col=2):
    row[0].fill = head_fill
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 60

# ===== Sheet 2: 技能 =====
ws = wb.create_sheet("技能")
skills = [
    ["需求分析", "90%", "精通"],
    ["产品规划", "85%", "熟练"],
    ["用户研究", "80%", "熟练"],
    ["原型设计（Axure / Figma / 墨刀）", "75%", "熟练"],
    ["数据分析（SQL）", "70%", "熟悉"],
    ["项目管理（敏捷）", "75%", "熟练"],
]
style_title(ws, 3, "专业技能")
heads = ["技能", "熟练度", "掌握程度"]
for i, h in enumerate(heads, 1):
    c = ws.cell(2, i, h)
    c.font = head_font
    c.fill = head_fill
    c.alignment = center
    c.border = border
for r, row in enumerate(skills, start=3):
    for i, val in enumerate(row, 1):
        write_cell(ws, r, i, val, alignment=center if i > 1 else left)
ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12

# ===== Sheet 3: 工作经历 =====
ws = wb.create_sheet("工作经历")
style_title(ws, 4, "工作经历")
heads = ["时间段", "公司", "职位", "主要职责"]
for i, h in enumerate(heads, 1):
    c = ws.cell(2, i, h)
    c.font = head_font
    c.fill = head_fill
    c.alignment = center
    c.border = border
exp = [
    ["2023 – 至今", "某某互联网公司", "产品经理",
     "负责企业级管理后台的需求梳理与迭代规划，推动 3 个大版本上线；通过用户调研与数据分析优化核心流程，关键指标提升 25%。"],
    ["2022 – 2023", "某某科技公司", "产品助理",
     "协助完成社区 App 的原型设计与 PRD 撰写，跟进开发与验收；参与用户反馈收集与竞品分析，输出多份产品调研报告。"],
]
for r, row in enumerate(exp, start=3):
    for i, val in enumerate(row, 1):
        write_cell(ws, r, i, val, alignment=center if i <= 3 else left)
widths = [14, 18, 14, 60]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 4: 项目作品 =====
ws = wb.create_sheet("项目作品")
style_title(ws, 4, "项目作品")
heads = ["项目名称", "简介", "关键词/工具", "链接"]
for i, h in enumerate(heads, 1):
    c = ws.cell(2, i, h)
    c.font = head_font
    c.fill = head_fill
    c.alignment = center
    c.border = border
projects = [
    ["B 端 SaaS 后台改版", "企业级管理后台从 0 到 1，负责需求梳理、信息架构与原型设计。",
     "B 端产品 / Axure", ""],
    ["移动端社区 App", "负责用户增长与留存优化，通过 A/B 测试迭代核心体验。",
     "用户增长 / A/B 测试", ""],
    ["数据看板产品", "面向运营团队的数据可视化平台，沉淀指标体系与报表能力。",
     "数据产品 / Figma", ""],
    ["AI越主动体验就越好吗？（文章）", "人人都是产品经理专栏文章，探讨 AI 产品主动性与用户体验的边界。",
     "产品思考 / 写作", "https://www.woshipm.com/evaluating/6443589.html"],
]
for r, row in enumerate(projects, start=3):
    for i, val in enumerate(row, 1):
        write_cell(ws, r, i, val, alignment=left if i == 2 else center)
widths = [26, 42, 22, 46]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 冻结表头
for ws in wb.worksheets[1:]:
    ws.freeze_panes = "A3"
ws = wb["基本信息"]
ws.freeze_panes = "A2"

wb.save("个人信息.xlsx")
print("已生成：个人信息.xlsx")
